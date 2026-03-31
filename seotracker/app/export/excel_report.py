"""
Excel/XLSX report generator for crawl results.

Generates a multi-sheet workbook with:
  - Summary overview
  - All pages (key SEO fields)
  - Issues by severity
  - Redirect report
  - Indexability report
  - Link graph summary
  - JS parity report
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Severity colors
SEVERITY_FILLS = {
    "critical": PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid"),
    "high": PatternFill(start_color="FF8800", end_color="FF8800", fill_type="solid"),
    "medium": PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid"),
    "low": PatternFill(start_color="88CCFF", end_color="88CCFF", fill_type="solid"),
}

HEADER_FILL = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _style_header_row(ws, col_count: int):
    """Apply header styling to the first row."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _auto_width(ws, min_width: int = 10, max_width: int = 60):
    """Auto-fit column widths."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells[:100]:  # sample first 100 rows
            val = str(cell.value or "")
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))


def generate_excel_report(
    pages: list[dict],
    links: list[dict],
    issues: list[dict],
    summary: dict,
    output_path: str | Path,
) -> Path:
    """
    Generate a comprehensive Excel report from crawl + audit results.

    Returns the path to the saved file.
    """
    wb = Workbook()
    output_path = Path(output_path)

    # --- Sheet 1: Summary ---
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _build_summary_sheet(ws_summary, pages, issues, summary)

    # --- Sheet 2: Issues ---
    ws_issues = wb.create_sheet("Issues")
    _build_issues_sheet(ws_issues, issues)

    # --- Sheet 3: All Pages ---
    ws_pages = wb.create_sheet("Pages")
    _build_pages_sheet(ws_pages, pages)

    # --- Sheet 4: Redirects ---
    ws_redirects = wb.create_sheet("Redirects")
    _build_redirects_sheet(ws_redirects, pages)

    # --- Sheet 5: Indexability ---
    ws_index = wb.create_sheet("Indexability")
    _build_indexability_sheet(ws_index, pages)

    # --- Sheet 6: Link Graph ---
    ws_links = wb.create_sheet("Link Graph")
    _build_link_graph_sheet(ws_links, pages, links)

    # --- Sheet 7: JS Parity ---
    ws_js = wb.create_sheet("JS Parity")
    _build_js_parity_sheet(ws_js, pages)

    # Save
    wb.save(output_path)
    logger.info("Excel report saved: %s", output_path)
    return output_path


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_summary_sheet(ws, pages: list[dict], issues: list[dict], summary: dict):
    """Overview sheet with key metrics."""
    stats = summary.get("stats", {})
    issue_summary = summary.get("issue_summary", {})

    # Title
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = f"SEO Audit Report — {summary.get('domain', 'Unknown')}"
    title_cell.font = Font(bold=True, size=16, color="2B5797")

    ws["A2"].value = f"Generated: {summary.get('timestamp', datetime.now().isoformat())}"
    ws["A2"].font = Font(italic=True, color="666666")

    # Crawl stats
    row = 4
    ws.cell(row=row, column=1, value="CRAWL STATISTICS").font = Font(bold=True, size=13)
    row += 1
    stat_items = [
        ("Pages Crawled", stats.get("pages_crawled", len(pages))),
        ("Pages Rendered (JS)", stats.get("pages_rendered", 0)),
        ("Links Found", stats.get("links_found", 0)),
        ("Errors", stats.get("errors", 0)),
        ("Elapsed (seconds)", stats.get("elapsed_seconds", 0)),
        ("Pages/second", stats.get("pages_per_second", 0)),
        ("Sitemap URLs", summary.get("sitemap_urls_count", 0)),
    ]
    for label, value in stat_items:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1

    # Issue summary
    row += 1
    ws.cell(row=row, column=1, value="ISSUE SUMMARY").font = Font(bold=True, size=13)
    row += 1
    severity_items = [
        ("Critical", issue_summary.get("critical", 0), "FF4444"),
        ("High", issue_summary.get("high", 0), "FF8800"),
        ("Medium", issue_summary.get("medium", 0), "FFCC00"),
        ("Low", issue_summary.get("low", 0), "88CCFF"),
        ("Total", sum(issue_summary.values()), "2B5797"),
    ]
    for label, count, color in severity_items:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, color=color)
        ws.cell(row=row, column=2, value=count)
        row += 1

    # Indexability breakdown
    row += 1
    ws.cell(row=row, column=1, value="INDEXABILITY").font = Font(bold=True, size=13)
    row += 1
    indexable = sum(1 for p in pages if p.get("is_indexable", True) and (p.get("status_code") or 0) == 200)
    noindex = sum(1 for p in pages if p.get("is_noindex"))
    blocked = sum(1 for p in pages if not p.get("robots_txt_allowed", True))
    errors = sum(1 for p in pages if (p.get("status_code") or 0) >= 400)

    for label, count in [
        ("Indexable pages", indexable),
        ("Noindex pages", noindex),
        ("Robots blocked", blocked),
        ("Error pages (4xx/5xx)", errors),
    ]:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=count)
        row += 1

    _auto_width(ws)


def _build_issues_sheet(ws, issues: list[dict]):
    """All issues sorted by severity."""
    headers = [
        "Severity", "Issue Type", "Confidence", "Affected URLs",
        "Example URL", "How to Fix", "Why It Matters",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, len(headers))

    # Sort: critical > high > medium > low
    severity_order = {"critical": 0, "high": 1, "medium": 2, "low": 3}
    sorted_issues = sorted(issues, key=lambda i: severity_order.get(i.get("severity", "low"), 9))

    for row_idx, issue in enumerate(sorted_issues, 2):
        sev = issue.get("severity", "low")
        ws.cell(row=row_idx, column=1, value=sev.upper())
        ws.cell(row=row_idx, column=1).fill = SEVERITY_FILLS.get(sev, PatternFill())
        ws.cell(row=row_idx, column=1).font = Font(bold=True, color="FFFFFF" if sev in ("critical", "high") else "000000")
        ws.cell(row=row_idx, column=2, value=issue.get("issue_type", ""))
        ws.cell(row=row_idx, column=3, value=issue.get("confidence", 1.0))
        ws.cell(row=row_idx, column=4, value=issue.get("affected_urls_count", 0))
        ws.cell(row=row_idx, column=5, value=issue.get("affected_url", ""))
        ws.cell(row=row_idx, column=6, value=issue.get("how_to_fix", ""))
        ws.cell(row=row_idx, column=7, value=issue.get("why_it_matters", ""))

    ws.auto_filter.ref = ws.dimensions
    _auto_width(ws)


def _build_pages_sheet(ws, pages: list[dict]):
    """All pages with key SEO fields."""
    headers = [
        "URL", "Status", "Title", "Title Len", "Meta Desc Len",
        "H1", "H1 Count", "Word Count", "Canonical",
        "Indexable", "Reason", "Depth",
        "Int Links", "Ext Links", "Imgs", "Missing Alt",
        "TTFB (ms)", "Size (bytes)", "Redirects", "Rendered",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, len(headers))

    for row_idx, p in enumerate(pages, 2):
        vals = [
            p.get("url", ""),
            p.get("status_code", ""),
            p.get("title", ""),
            p.get("title_length", ""),
            p.get("meta_description_length", ""),
            p.get("h1_text", ""),
            p.get("h1_count", 0),
            p.get("word_count", 0),
            p.get("canonical_url", ""),
            "Yes" if p.get("is_indexable", True) else "No",
            p.get("indexability_reason", ""),
            p.get("depth", 0),
            p.get("internal_links_count", 0),
            p.get("external_links_count", 0),
            p.get("img_count", 0),
            p.get("img_missing_alt", 0),
            round(p.get("ttfb_ms") or 0, 1),
            p.get("response_bytes", 0),
            p.get("redirect_hops", 0),
            "Yes" if p.get("was_rendered") else "No",
        ]
        for col, val in enumerate(vals, 1):
            ws.cell(row=row_idx, column=col, value=val)

        # Color status codes
        status = p.get("status_code") or 0
        if status >= 500:
            ws.cell(row=row_idx, column=2).fill = SEVERITY_FILLS["critical"]
        elif status >= 400:
            ws.cell(row=row_idx, column=2).fill = SEVERITY_FILLS["high"]
        elif 300 <= status < 400:
            ws.cell(row=row_idx, column=2).fill = SEVERITY_FILLS["medium"]

    ws.auto_filter.ref = ws.dimensions
    _auto_width(ws)


def _build_redirects_sheet(ws, pages: list[dict]):
    """Pages with redirect chains."""
    headers = ["URL", "Final URL", "Hops", "Chain", "Status"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, len(headers))

    row_idx = 2
    for p in pages:
        chain = p.get("redirect_chain") or []
        if not chain:
            continue
        chain_str = " → ".join(
            f"{hop.get('url', '?')} ({hop.get('status_code', '?')})" for hop in chain
        )
        ws.cell(row=row_idx, column=1, value=p.get("url", ""))
        ws.cell(row=row_idx, column=2, value=p.get("final_url", ""))
        ws.cell(row=row_idx, column=3, value=len(chain))
        ws.cell(row=row_idx, column=4, value=chain_str)
        ws.cell(row=row_idx, column=5, value=p.get("status_code", ""))
        row_idx += 1

    if row_idx == 2:
        ws.cell(row=2, column=1, value="No redirects detected")

    ws.auto_filter.ref = ws.dimensions
    _auto_width(ws)


def _build_indexability_sheet(ws, pages: list[dict]):
    """Indexability status for all pages."""
    headers = [
        "URL", "Status", "Indexable", "Reason",
        "Meta Robots", "Canonical", "Robots Allowed", "Noindex",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, len(headers))

    for row_idx, p in enumerate(pages, 2):
        ws.cell(row=row_idx, column=1, value=p.get("url", ""))
        ws.cell(row=row_idx, column=2, value=p.get("status_code", ""))
        ws.cell(row=row_idx, column=3, value="Yes" if p.get("is_indexable", True) else "No")
        ws.cell(row=row_idx, column=4, value=p.get("indexability_reason", ""))
        ws.cell(row=row_idx, column=5, value=p.get("meta_robots", ""))
        ws.cell(row=row_idx, column=6, value=p.get("canonical_url", ""))
        ws.cell(row=row_idx, column=7, value="Yes" if p.get("robots_txt_allowed", True) else "No")
        ws.cell(row=row_idx, column=8, value="Yes" if p.get("is_noindex") else "No")

        if not p.get("is_indexable", True):
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = PatternFill(
                    start_color="FFE0E0", end_color="FFE0E0", fill_type="solid"
                )

    ws.auto_filter.ref = ws.dimensions
    _auto_width(ws)


def _build_link_graph_sheet(ws, pages: list[dict], links: list[dict]):
    """Link graph summary: inlinks/outlinks per page."""
    from collections import Counter

    inlinks = Counter()
    outlinks = Counter()
    for link in links:
        if link.get("is_internal"):
            outlinks[link.get("source_url", "")] += 1
            inlinks[link.get("dest_url", link.get("dest_url_normalized", ""))] += 1

    headers = ["URL", "Depth", "Inlinks", "Outlinks", "Indexable"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, len(headers))

    # Sort by inlinks descending
    page_data = []
    for p in pages:
        url = p.get("url", "")
        page_data.append({
            "url": url,
            "depth": p.get("depth", 0),
            "inlinks": inlinks.get(url, 0) + inlinks.get(p.get("url_normalized", ""), 0),
            "outlinks": outlinks.get(url, 0),
            "indexable": p.get("is_indexable", True),
        })
    page_data.sort(key=lambda x: x["inlinks"], reverse=True)

    for row_idx, pd in enumerate(page_data, 2):
        ws.cell(row=row_idx, column=1, value=pd["url"])
        ws.cell(row=row_idx, column=2, value=pd["depth"])
        ws.cell(row=row_idx, column=3, value=pd["inlinks"])
        ws.cell(row=row_idx, column=4, value=pd["outlinks"])
        ws.cell(row=row_idx, column=5, value="Yes" if pd["indexable"] else "No")

    ws.auto_filter.ref = ws.dimensions
    _auto_width(ws)


def _build_js_parity_sheet(ws, pages: list[dict]):
    """JS render parity report."""
    headers = [
        "URL", "Title Changed", "H1 Changed", "Canonical Changed",
        "Meta Robots Changed", "Links Added", "Word Count Raw",
        "Word Count Rendered", "Console Errors", "Significant Diff",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, len(headers))

    row_idx = 2
    for p in pages:
        if not p.get("was_rendered"):
            continue
        diff = p.get("raw_vs_rendered_diff") or {}
        ws.cell(row=row_idx, column=1, value=p.get("url", ""))
        ws.cell(row=row_idx, column=2, value="Yes" if diff.get("title_changed") else "No")
        ws.cell(row=row_idx, column=3, value="Yes" if diff.get("h1_changed") else "No")
        ws.cell(row=row_idx, column=4, value="Yes" if diff.get("canonical_changed") else "No")
        ws.cell(row=row_idx, column=5, value="Yes" if diff.get("meta_robots_changed") else "No")
        ws.cell(row=row_idx, column=6, value=diff.get("links_added", 0))
        ws.cell(row=row_idx, column=7, value=diff.get("word_count_raw", 0))
        ws.cell(row=row_idx, column=8, value=diff.get("word_count_rendered", 0))
        ws.cell(row=row_idx, column=9, value=len(p.get("console_errors") or []))
        ws.cell(row=row_idx, column=10, value="Yes" if diff.get("has_significant_diff") else "No")

        if diff.get("has_significant_diff"):
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = PatternFill(
                    start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"
                )
        row_idx += 1

    if row_idx == 2:
        ws.cell(row=2, column=1, value="No pages were JS-rendered in this run")

    ws.auto_filter.ref = ws.dimensions
    _auto_width(ws)
